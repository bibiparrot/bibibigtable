import traceback
from pstats import SortKey

import python_calamine  # necessary for pandas
import pandas as pd
from loguru import logger


def to_color_excel_openpyxl(df_data: pd.DataFrame, xlsx_path, write_index=True):
    """
    colorize pandas.DataFrame according to quartile ratio
    ...

    Args
    ----------
    df_data : pandas.DataFrame
        data frame
    xlsx_path : str
        saving path of excel

    """
    from openpyxl.styles import PatternFill
    from openpyxl.workbook import Workbook
    score_titles = ["score", "scores", "value", "values", "打分", "分值", "分数", "评分"]
    try:
        if write_index:
            df_data = df_data.copy(deep=True).reset_index()
        else:
            df_data = df_data.copy(deep=True).reset_index(drop=True)
        wb = Workbook()
        ws = wb.active

        for c_idx, column in enumerate(df_data.columns):
            ws.cell(row=1, column=c_idx + 1, value=column)

        for r_idx, row in df_data.iterrows():
            for c_idx, val in enumerate(row):
                try:
                    ws.cell(row=r_idx + 2, column=c_idx + 1, value=val)
                except:
                    ws.cell(row=r_idx + 2, column=c_idx + 1, value=str(val))
        quartiles = [0, 0.25, 0.5, 0.75, 1]
        colors_light = ["00ffd4", "98f5ff", "fed8b1", "ff7f7d"]
        colors_dark = ["40E0D0", "6495ED", "FF7F50", "FC3468"]
        for c_idx, column in enumerate(df_data.columns):
            try:
                pd.to_numeric(df_data[column])
                colors = colors_light
            except:
                continue
            if any(subs in str(column).lower() for subs in score_titles):
                colors = colors_dark
            values = pd.to_numeric(df_data[column])
            quartile_values = [values.quantile(q) for q in quartiles]
            for r_idx, val in enumerate(values):
                cell = ws.cell(row=r_idx + 2, column=c_idx + 1)
                for i in range(len(quartiles) - 1):
                    if val <= quartile_values[i + 1]:
                        cell.fill = PatternFill(
                            start_color=colors[i],
                            end_color=colors[i],
                            fill_type="solid",
                        )
                        break
        wb.save(xlsx_path)
    except Exception as ex:
        logger.warning(
            f"to_color_excel() FAILED because of {ex}, using pandas DataFrame to_excel()"
        )
        logger.warning(traceback.print_exc())
        df_data.to_excel(xlsx_path)


def to_color_excel_xlsxwriter(df_data: pd.DataFrame, xlsx_path, write_index=True):
    """
    colorize pandas.DataFrame according to quantile ratio

    for row in range(0, row_max):
        for col in range(0, col_max):
            worksheet.write(row, col, some_data)
    ...

    Args
    ----------
    df_data : pandas.DataFrame
        data frame
    xlsx_path : str
        saving path of excel



    """
    import xlsxwriter, io
    score_titles = ["score", "scores", "value", "values", "打分", "分值", "分数", "评分"]
    try:
        if write_index:
            df_data = df_data.copy(deep=True).reset_index()
        else:
            df_data = df_data.copy(deep=True).reset_index(drop=True)
        workbook = xlsxwriter.Workbook(xlsx_path)
        worksheet = workbook.add_worksheet()

        # prepare format
        bold_format = workbook.add_format({'bold': True})
        colors_light = ["00ffd4", "98f5ff", "fed8b1", "ff7f7d"]
        colors_dark = ["40E0D0", "6495ED", "FF7F50", "FC3468"]
        datetime_format = workbook.add_format({'num_format': 'yyyy-MM-dd HH:mm:ss'})
        colors_light_formats = [workbook.add_format({'bg_color': color}) for color in colors_light]
        colors_dark_formats = [workbook.add_format({'bg_color': color}) for color in colors_dark]

        # prepare quantile values and color
        quantiles = [0, 0.25, 0.5, 0.75, 1]
        column_quantile_values = {}
        column_colors = {}
        for c_idx, column in enumerate(df_data.columns):
            try:
                values = pd.to_numeric(df_data[column])
                colors = colors_light_formats
            except:
                continue
            if any(subs in str(column).lower() for subs in score_titles):
                colors = colors_dark_formats
            quantile_values = [values.quantile(q) for q in quantiles]
            column_quantile_values[c_idx] = quantile_values
            column_colors[c_idx] = colors

        # write head
        for c_idx, column in enumerate(df_data.columns):
            worksheet.write_string(0, c_idx, column, bold_format)
        dtypes = df_data.dtypes
        # write datas
        for r_idx, row in df_data.iterrows():
            for c_idx, val in enumerate(row):
                cell_format = None
                quantile_values = column_quantile_values.get(c_idx, [])
                if len(quantile_values) > 0:
                    v = pd.to_numeric(val, errors='coerce')
                    for i in range(len(quantiles) - 1):
                        if v <= quantile_values[i + 1]:
                            cell_format = column_colors[c_idx][i]
                            break
                try:
                    if pd.api.types.is_datetime64_any_dtype(dtypes.iloc[c_idx]):
                        if cell_format is None:
                            cell_format = datetime_format
                        worksheet.write_datetime(r_idx + 1, c_idx, val, cell_format)
                    else:
                        worksheet.write(r_idx + 1, c_idx, val, cell_format)
                except Exception as exp:
                    logger.warning(
                        f"worksheet.write() FAILED because of {exp}"
                    )
                    worksheet.write(r_idx + 1, c_idx, str(val), cell_format)
        workbook.close()
    except Exception as ex:
        logger.warning(
            f"to_color_excel() FAILED because of {ex}, using pandas DataFrame to_excel()"
        )
        logger.warning(traceback.print_exc())
        df_data.to_excel(xlsx_path)


def large_excel_to_csv(io_excel, csv_path_or_buf, sheet_name=None, chunk_size=None, encoding=None):
    '''
    read very large excel to csv iteratively
    :param io_excel: excel file
    :param csv_path_or_buf: csv file
    :param sheet_name: excel sheet name
    :param chunk_size: chunk size for iteratively reading
    :param encoding: csv encoding
    :return:
    '''
    if chunk_size is None:
        chunk_size = 1024 * 4
    if encoding is None:
        encoding = 'utf8'
    reader = pd.read_excel(io_excel, sheet_name=sheet_name, chunksize=chunk_size, engine="calamine")
    for i, chunk in enumerate(reader):
        if i == 0:
            chunk.to_csv(csv_path_or_buf, encoding=encoding, index=False, mode='w')
        else:
            chunk.to_csv(csv_path_or_buf, encoding=encoding, index=False, header=False, mode='a')
    return csv_path_or_buf


def read_large_excel_calamine(io, *args, **kwargs):
    '''
    see arguments in pandas.read_excel()
    :param io:
    :param args:
    :param kwargs:
    :return: pandas.DataFrame
    '''
    calamine_df = pd.read_excel(io, engine="calamine", *args, **kwargs)
    return calamine_df


def read_large_excel_openpyxl(io, sheet_name="Sheet1"):
    import openpyxl
    workbook = openpyxl.load_workbook(io, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = []
    for row in worksheet.iter_rows(values_only=True):
        rows.append(row)
    dataDF = pd.DataFrame(rows[1:], columns=rows[0])
    return dataDF


def read_sql_to_hdf(sql, con, hdf_file, chunk_size=None, encoding=None, complevel=None):
    '''
    :param sql: sql in string
    :param con: ADBC Connection, SQLAlchemy connectable, str, or sqlite3 connection
        ADBC provides high performance I/O with native type support, where available.
        Using SQLAlchemy makes it possible to use any DB supported by that
        library. If a DBAPI2 object, only sqlite3 is supported. The user is responsible
        for engine disposal and connection closure for the ADBC connection and
        SQLAlchemy connectable; str connections are closed automatically. See
        `here <https://docs.sqlalchemy.org/en/20/core/connections.html>`_.
        for example:
            import cx_Oracle
            dsn_tns = cx_Oracle.makedsn('ORACLE_HOSTNAME', 'PORT', service_name='SERVICE_NAME')
            db = cx_Oracle.connect(user='USERNAME', password='PASSWORD', dsn=dsn_tns)
    :param hdf_file: hdf file
    :param hdf_key: hdf key
    :param chunk_size: chunk size
    :return: hdf_file
    '''
    if chunk_size is None:
        chunk_size = 1024 * 4
    if encoding is None:
        encoding = 'utf8'
    iterator = pd.read_sql(sql, con, chunksize=chunk_size)

    for i, chunk in enumerate(iterator):
        logger.info(f"Processing chunk {i}")
        if i == 0:
            chunk.to_hdf(hdf_file, key=hdf_file, mode='w', append=False, complevel=complevel, encoding=encoding)
        else:
            chunk.to_hdf(hdf_file, key=hdf_file, mode='a', append=True, complevel=complevel, encoding=encoding)
    return hdf_file


def read_sql_to_csv(sql, con, csv_file, chunk_size=None, encoding=None):
    '''
    :param sql: sql in string
    :param con: ADBC Connection, SQLAlchemy connectable, str, or sqlite3 connection
        ADBC provides high performance I/O with native type support, where available.
        Using SQLAlchemy makes it possible to use any DB supported by that
        library. If a DBAPI2 object, only sqlite3 is supported. The user is responsible
        for engine disposal and connection closure for the ADBC connection and
        SQLAlchemy connectable; str connections are closed automatically. See
        `here <https://docs.sqlalchemy.org/en/20/core/connections.html>`_.
        for example:
            import cx_Oracle
            dsn_tns = cx_Oracle.makedsn('ORACLE_HOSTNAME', 'PORT', service_name='SERVICE_NAME')
            db = cx_Oracle.connect(user='USERNAME', password='PASSWORD', dsn=dsn_tns)
    :param csv_file: csv file
    :param chunk_size: chunk size
    :param encoding: csv encoding
    :return: csv file
    '''
    if chunk_size is None:
        chunk_size = 1024 * 4
    if encoding is None:
        encoding = 'utf8'
    iterator = pd.read_sql(sql, con=con, chunksize=chunk_size)

    for i, chunk in enumerate(iterator):
        logger.info(f"Processing chunk {i}")
        if i == 0:
            chunk.to_csv(csv_file, mode='w', encoding=encoding, header=True, index=False)
        else:
            chunk.to_csv(csv_file, mode='a', encoding=encoding, header=False, index=False)
    return csv_file


########################################################################################################################


if __name__ == "__main__":
    import numpy as np

    import pandas as pd
    import numpy as np
    import random
    from datetime import datetime, timedelta

    # Define the number of rows for the DataFrame
    num_rows = 10000

    # Generate random datetime values
    start_date = datetime(2020, 1, 1)
    end_date = datetime(2024, 1, 1)
    date_range = [start_date + timedelta(days=random.randint(0, (end_date - start_date).days)) for _ in range(num_rows)]

    # Generate random text values
    text_values = [''.join(random.choices('abcdefghijklmnopqrstuvwxyz', k=random.randint(5, 15))) for _ in
                   range(num_rows)]

    # Define categories
    categories = ['A', 'B', 'C']

    # Generate random category values
    category_values = [random.choice(categories) for _ in range(num_rows)]

    # Generate random float values
    float_values = [random.uniform(0, 100) for _ in range(num_rows)]

    # Generate random integer values
    int_values = [random.randint(1, 100) for _ in range(num_rows)]

    # Create the DataFrame
    df = pd.DataFrame({
        'Datetime': date_range,
        'Text': text_values,
        'Category': category_values,
        'Float': float_values,
        'Integer': int_values
    })
    import cProfile

    with cProfile.Profile() as pr:
        logger.info("to_color_excel_openpyxl start...")
        to_color_excel_openpyxl(df, 'random_openpyxl.xlsx', write_index=True)
        logger.info("to_color_excel_openpyxl finished")

        logger.info("to_color_excel_xlsxwriter start...")
        to_color_excel_xlsxwriter(df, 'random_xlsxwriter.xlsx', write_index=True)
        logger.info("to_color_excel_xlsxwriter finished")

        pr.print_stats(sort=SortKey.CUMULATIVE)
